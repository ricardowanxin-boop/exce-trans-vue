from __future__ import annotations

import base64
import io
import json
import os
import re
import uuid
from pathlib import Path
from typing import Dict, Iterable, List, Tuple
from urllib.parse import quote

import httpx
import openpyxl
from fastapi import FastAPI, HTTPException
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import FileResponse
from pydantic import BaseModel

from core.excel_parser import apply_translations, extract_shape_entries, extract_texts, get_sheet_names


APP_DIR = Path(__file__).resolve().parent
ROOT_DIR = APP_DIR.parent
DEFAULT_RUNTIME_DIR = APP_DIR / ".runtime"
if (ROOT_DIR / "python_backend").exists():
    DEFAULT_RUNTIME_DIR = ROOT_DIR / "python_backend" / ".runtime"
RUNTIME_DIR = Path(os.getenv("PY_RUNTIME_DIR", str(DEFAULT_RUNTIME_DIR)))
UPLOAD_DIR = RUNTIME_DIR / "uploads"
RESULT_DIR = RUNTIME_DIR / "results"
META_DIR = RUNTIME_DIR / "meta"
API_PREFIX = os.getenv("PY_API_PREFIX", "/python-api")
PREVIEW_LIMIT_PER_SHEET = 30
TRANSLATE_BATCH_LIMIT = 80
TRANSLATE_CHAR_LIMIT = 12000
OPENAI_API_KEY = os.getenv("OPENAI_API_KEY", "")
OPENAI_BASE_URL = os.getenv("OPENAI_BASE_URL", "https://api-inference.modelscope.cn/v1/chat/completions")
MODEL_NAME = os.getenv("MODEL_NAME", "Qwen/Qwen3-30B-A3B-Instruct-2507")
MOCK_TRANSLATION = os.getenv("PY_MOCK_TRANSLATION", "false").lower() == "true"

for directory in (UPLOAD_DIR, RESULT_DIR, META_DIR):
    directory.mkdir(parents=True, exist_ok=True)


app = FastAPI(title="Excel Translator Python API")
app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)


class UploadInitRequest(BaseModel):
    action: str
    file_name: str = ""
    file_size: int = 0
    content_type: str = "application/octet-stream"


class UploadAppendRequest(BaseModel):
    action: str
    upload_id: str
    chunk_index: int
    chunk_base64: str


class UploadCompleteRequest(BaseModel):
    action: str
    upload_id: str
    total_chunks: int


class ParseRequest(BaseModel):
    user_id: str | None = None
    upload_id: str | None = None
    sheet_name: str | None = None


class TranslateRequest(BaseModel):
    user_id: str | None = None
    upload_id: str | None = None
    target_lang: str
    sheet_name: str | None = None
    sheet_names: List[str] | None = None


def success(data: Dict) -> Dict:
    return {"isBase64Encoded": False, "statusCode": 200, "data": data}


def sanitize_filename(name: str) -> str:
    clean = re.sub(r"[^\w.\-()\u4e00-\u9fff]+", "_", name or "translated.xlsx").strip("._")
    return clean or "translated.xlsx"


def meta_path(upload_id: str) -> Path:
    return META_DIR / f"{upload_id}.json"


def upload_path(upload_id: str) -> Path:
    return UPLOAD_DIR / f"{upload_id}.bin"


def load_meta(upload_id: str) -> Dict:
    path = meta_path(upload_id)
    if not path.exists():
        raise HTTPException(status_code=404, detail="上传会话不存在")
    return json.loads(path.read_text(encoding="utf-8"))


def save_meta(upload_id: str, payload: Dict) -> None:
    meta_path(upload_id).write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8")


def init_upload_session(file_name: str, file_size: int, content_type: str) -> str:
    upload_id = f"upload_{uuid.uuid4().hex}"
    payload = {
        "upload_id": upload_id,
        "file_name": sanitize_filename(file_name),
        "file_size": max(int(file_size or 0), 0),
        "content_type": content_type or "application/octet-stream",
        "status": "uploading",
        "next_chunk_index": 0,
        "received_chunks": 0,
        "total_chunks": 0,
    }
    upload_path(upload_id).write_bytes(b"")
    save_meta(upload_id, payload)
    return upload_id


def append_upload_chunk(upload_id: str, chunk_index: int, chunk_base64: str) -> None:
    meta = load_meta(upload_id)
    if meta.get("status") != "uploading":
        raise HTTPException(status_code=400, detail="上传会话已关闭")
    expected_index = int(meta.get("next_chunk_index", 0))
    if chunk_index != expected_index:
        raise HTTPException(status_code=400, detail=f"分片顺序错误，期望 {expected_index}，收到 {chunk_index}")

    chunk_bytes = base64.b64decode(chunk_base64)
    with upload_path(upload_id).open("ab") as file:
      file.write(chunk_bytes)

    meta["next_chunk_index"] = expected_index + 1
    meta["received_chunks"] = int(meta.get("received_chunks", 0)) + 1
    save_meta(upload_id, meta)


def complete_upload_session(upload_id: str, total_chunks: int) -> None:
    meta = load_meta(upload_id)
    if int(meta.get("received_chunks", 0)) != total_chunks:
        raise HTTPException(status_code=400, detail="上传分片数量不完整")
    meta["status"] = "completed"
    meta["total_chunks"] = total_chunks
    save_meta(upload_id, meta)


def get_file_bytes(upload_id: str) -> Tuple[bytes, Dict]:
    meta = load_meta(upload_id)
    if meta.get("status") != "completed":
        raise HTTPException(status_code=400, detail="文件仍在上传中，请稍后重试")
    path = upload_path(upload_id)
    if not path.exists():
        raise HTTPException(status_code=404, detail="上传文件不存在")
    return path.read_bytes(), meta


def build_sheet_previews(file_bytes: bytes) -> Tuple[List[str], Dict[str, List[Dict[str, str]]], Dict[str, int]]:
    workbook = openpyxl.load_workbook(io.BytesIO(file_bytes), read_only=True, data_only=False)
    sheet_names = list(workbook.sheetnames)
    previews_by_sheet: Dict[str, List[Dict[str, str]]] = {}
    total_cells_by_sheet: Dict[str, int] = {}

    for sheet_name in sheet_names:
        worksheet = workbook[sheet_name]
        shape_preview = [
            {"coordinate": coordinate, "text": text, "source": "shape"}
            for coordinate, text in extract_shape_entries(file_bytes, sheet_name)
        ]
        preview: List[Dict[str, str]] = shape_preview[:PREVIEW_LIMIT_PER_SHEET]
        total = len(shape_preview)

        for row in worksheet.iter_rows():
            for cell in row:
                value = cell.value
                if value is None:
                    continue
                if cell.data_type == "f":
                    continue
                if isinstance(value, (int, float)):
                    continue
                text = str(value).strip()
                if not text:
                    continue
                total += 1
                if len(preview) < PREVIEW_LIMIT_PER_SHEET:
                    preview.append({"coordinate": cell.coordinate, "text": text, "source": "cell"})

        previews_by_sheet[sheet_name] = preview
        total_cells_by_sheet[sheet_name] = total

    workbook.close()
    return sheet_names, previews_by_sheet, total_cells_by_sheet


def chunk_items(items: Dict[str, str]) -> Iterable[Dict[str, str]]:
    batch: Dict[str, str] = {}
    total_chars = 0
    for key, value in items.items():
        value_chars = len(value)
        if batch and (len(batch) >= TRANSLATE_BATCH_LIMIT or total_chars + value_chars > TRANSLATE_CHAR_LIMIT):
            yield batch
            batch = {}
            total_chars = 0
        batch[key] = value
        total_chars += value_chars
    if batch:
        yield batch


async def call_llm_translate(items: Dict[str, str], target_lang: str) -> Dict[str, str]:
    if not items:
        return {}
    if MOCK_TRANSLATION:
        return {key: f"[{target_lang}] {value}" for key, value in items.items()}
    if not OPENAI_API_KEY:
        raise RuntimeError("缺少 OPENAI_API_KEY")

    system_prompt = (
        f"你是一个专业的翻译专家。请将用户提供的 JSON 格式文本翻译成 {target_lang}。"
        "保持 JSON 的键名不变，只翻译值。返回必须是合法 JSON 对象。"
    )

    async with httpx.AsyncClient(timeout=120) as client:
        response = await client.post(
            OPENAI_BASE_URL,
            json={
                "model": MODEL_NAME,
                "messages": [
                    {"role": "system", "content": system_prompt},
                    {"role": "user", "content": json.dumps(items, ensure_ascii=False)},
                ],
                "temperature": 0.3,
                "response_format": {"type": "json_object"},
            },
            headers={
                "Content-Type": "application/json",
                "Authorization": f"Bearer {OPENAI_API_KEY}",
            },
        )
        response.raise_for_status()
        payload = response.json()
        content = payload["choices"][0]["message"]["content"]
        clean = content.replace("```json", "").replace("```", "").strip()
        parsed = json.loads(clean)
        return {key: str(value) for key, value in parsed.items()}


async def translate_sheet(file_bytes: bytes, sheet_name: str, target_lang: str) -> Tuple[bytes, int]:
    original_map = extract_texts(file_bytes, sheet_name)
    if not original_map:
        return file_bytes, 0

    merged_translations: Dict[str, str] = {}
    for batch in chunk_items(original_map):
        translated = await call_llm_translate(batch, target_lang)
        merged_translations.update(translated)

    updated_bytes = apply_translations(
        file_bytes=file_bytes,
        sheet_name=sheet_name,
        translations=merged_translations,
        original_texts_map=original_map,
    )
    return updated_bytes, len(original_map)


@app.get("/health")
async def health() -> Dict:
    return {"ok": True}


@app.post("/upload-file-function")
async def upload_file(payload: Dict) -> Dict:
    action = payload.get("action")

    if action == "init":
        body = UploadInitRequest(**payload)
        upload_id = init_upload_session(body.file_name, body.file_size, body.content_type)
        return success({"upload_id": upload_id})

    if action == "append":
        body = UploadAppendRequest(**payload)
        append_upload_chunk(body.upload_id, body.chunk_index, body.chunk_base64)
        return success({"ok": True})

    if action == "complete":
        body = UploadCompleteRequest(**payload)
        complete_upload_session(body.upload_id, body.total_chunks)
        return success({"upload_id": body.upload_id, "total_chunks": body.total_chunks})

    raise HTTPException(status_code=400, detail="不支持的操作类型")


@app.post("/parse-function")
async def parse_function(body: ParseRequest) -> Dict:
    if not body.upload_id:
        raise HTTPException(status_code=400, detail="缺少 upload_id")

    file_bytes, _meta = get_file_bytes(body.upload_id)
    sheet_names, previews_by_sheet, total_cells_by_sheet = build_sheet_previews(file_bytes)

    resolved_sheet = body.sheet_name if body.sheet_name in previews_by_sheet else (sheet_names[0] if sheet_names else "")
    total_cells = sum(total_cells_by_sheet.values())
    return success(
        {
            "preview": previews_by_sheet.get(resolved_sheet, []),
            "total_cells": total_cells,
            "sheet_name": resolved_sheet,
            "sheet_names": sheet_names,
            "previews_by_sheet": previews_by_sheet,
            "total_cells_by_sheet": total_cells_by_sheet,
        }
    )


@app.post("/translate-function")
async def translate_function(body: TranslateRequest) -> Dict:
    if not body.upload_id:
        raise HTTPException(status_code=400, detail="缺少 upload_id")

    file_bytes, meta = get_file_bytes(body.upload_id)
    sheet_names = body.sheet_names or ([body.sheet_name] if body.sheet_name else get_sheet_names(file_bytes)[:1])
    if not sheet_names:
        raise HTTPException(status_code=400, detail="未找到可翻译的工作表")

    used_count = 0
    current_bytes = file_bytes
    for sheet_name in sheet_names:
        current_bytes, translated_count = await translate_sheet(current_bytes, sheet_name, body.target_lang)
        used_count += 1 if translated_count > 0 else 0

    result_id = uuid.uuid4().hex
    original_name = sanitize_filename(meta.get("file_name") or "translated.xlsx")
    stem, suffix = os.path.splitext(original_name)
    suffix = suffix or ".xlsx"
    output_name = sanitize_filename(f"{stem}-translated{suffix}")
    result_path = RESULT_DIR / f"{result_id}{suffix}"
    result_path.write_bytes(current_bytes)

    return success(
        {
            "file_url": f"{API_PREFIX}/download/{result_id}?filename={quote(output_name)}",
            "used_count": used_count or 1,
        }
    )


@app.get("/download/{result_id}")
async def download_result(result_id: str, filename: str | None = None) -> FileResponse:
    files = list(RESULT_DIR.glob(f"{result_id}.*"))
    if not files:
        raise HTTPException(status_code=404, detail="结果文件不存在")
    path = files[0]
    download_name = sanitize_filename(filename or path.name)
    return FileResponse(path, filename=download_name, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
