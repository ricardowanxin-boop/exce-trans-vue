from __future__ import annotations

import base64
import io
import json
import os
import re
import uuid
from http.server import BaseHTTPRequestHandler, ThreadingHTTPServer
from pathlib import Path
from typing import Dict, Iterable, List, Tuple
from urllib.parse import parse_qs, quote, urlparse

import httpx
import openpyxl

from core.excel_parser import apply_translations, extract_shape_entries, extract_texts


APP_DIR = Path(__file__).resolve().parent
RUNTIME_DIR = Path(os.getenv("PY_RUNTIME_DIR", str(APP_DIR / ".runtime")))
UPLOAD_DIR = RUNTIME_DIR / "uploads"
RESULT_DIR = RUNTIME_DIR / "results"
META_DIR = RUNTIME_DIR / "meta"
API_PREFIX = os.getenv("PY_API_PREFIX", "/python-api")
PREVIEW_LIMIT_PER_SHEET = 30
TRANSLATE_BATCH_LIMIT = 80
TRANSLATE_CHAR_LIMIT = 12000
OPENAI_API_KEY = os.getenv("OPENAI_API_KEY", "")
OPENAI_BASE_URL = os.getenv("OPENAI_BASE_URL", "https://api-inference.modelscope.cn/v1/chat/completions")
MODEL_NAME = os.getenv("MODEL_NAME", "Qwen/Qwen3-30B-A3B-Instruct-2507")
MOCK_TRANSLATION = os.getenv("PY_MOCK_TRANSLATION", "false").lower() == "true"

for directory in (UPLOAD_DIR, RESULT_DIR, META_DIR):
    directory.mkdir(parents=True, exist_ok=True)


def sanitize_filename(name: str) -> str:
    clean = re.sub(r"[^\w.\-()\u4e00-\u9fff]+", "_", name or "translated.xlsx").strip("._")
    return clean or "translated.xlsx"


def meta_path(upload_id: str) -> Path:
    return META_DIR / f"{upload_id}.json"


def upload_path(upload_id: str) -> Path:
    return UPLOAD_DIR / f"{upload_id}.bin"


def load_meta(upload_id: str) -> Dict:
    path = meta_path(upload_id)
    if not path.exists():
        raise FileNotFoundError("上传会话不存在")
    return json.loads(path.read_text(encoding="utf-8"))


def save_meta(upload_id: str, payload: Dict) -> None:
    meta_path(upload_id).write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8")


def init_upload_session(file_name: str, file_size: int, content_type: str) -> str:
    upload_id = f"upload_{uuid.uuid4().hex}"
    payload = {
        "upload_id": upload_id,
        "file_name": sanitize_filename(file_name),
        "file_size": max(int(file_size or 0), 0),
        "content_type": content_type or "application/octet-stream",
        "status": "uploading",
        "next_chunk_index": 0,
        "received_chunks": 0,
        "total_chunks": 0,
    }
    upload_path(upload_id).write_bytes(b"")
    save_meta(upload_id, payload)
    return upload_id


def append_upload_chunk(upload_id: str, chunk_index: int, chunk_base64: str) -> None:
    meta = load_meta(upload_id)
    if meta.get("status") != "uploading":
        raise ValueError("上传会话已关闭")
    expected_index = int(meta.get("next_chunk_index", 0))
    if chunk_index != expected_index:
        raise ValueError(f"分片顺序错误，期望 {expected_index}，收到 {chunk_index}")

    chunk_bytes = base64.b64decode(chunk_base64)
    with upload_path(upload_id).open("ab") as file:
        file.write(chunk_bytes)

    meta["next_chunk_index"] = expected_index + 1
    meta["received_chunks"] = int(meta.get("received_chunks", 0)) + 1
    save_meta(upload_id, meta)


def complete_upload_session(upload_id: str, total_chunks: int) -> None:
    meta = load_meta(upload_id)
    if int(meta.get("received_chunks", 0)) != total_chunks:
        raise ValueError("上传分片数量不完整")
    meta["status"] = "completed"
    meta["total_chunks"] = total_chunks
    save_meta(upload_id, meta)


def get_file_bytes(upload_id: str) -> Tuple[bytes, Dict]:
    meta = load_meta(upload_id)
    if meta.get("status") != "completed":
        raise ValueError("文件仍在上传中，请稍后重试")
    path = upload_path(upload_id)
    if not path.exists():
        raise FileNotFoundError("上传文件不存在")
    return path.read_bytes(), meta


def build_sheet_previews(file_bytes: bytes) -> Tuple[List[str], Dict[str, List[Dict[str, str]]], Dict[str, int]]:
    workbook = openpyxl.load_workbook(io.BytesIO(file_bytes), read_only=True, data_only=False)
    sheet_names = list(workbook.sheetnames)
    previews_by_sheet: Dict[str, List[Dict[str, str]]] = {}
    total_cells_by_sheet: Dict[str, int] = {}

    for sheet_name in sheet_names:
        worksheet = workbook[sheet_name]
        shape_preview = [
            {"coordinate": coordinate, "text": text, "source": "shape"}
            for coordinate, text in extract_shape_entries(file_bytes, sheet_name)
        ]
        preview: List[Dict[str, str]] = shape_preview[:PREVIEW_LIMIT_PER_SHEET]
        total = len(shape_preview)

        for row in worksheet.iter_rows():
            for cell in row:
                value = cell.value
                if value is None:
                    continue
                if cell.data_type == "f":
                    continue
                if isinstance(value, (int, float)):
                    continue
                text = str(value).strip()
                if not text:
                    continue
                total += 1
                if len(preview) < PREVIEW_LIMIT_PER_SHEET:
                    preview.append({"coordinate": cell.coordinate, "text": text, "source": "cell"})

        previews_by_sheet[sheet_name] = preview
        total_cells_by_sheet[sheet_name] = total

    workbook.close()
    return sheet_names, previews_by_sheet, total_cells_by_sheet


def chunk_items(items: Dict[str, str]) -> Iterable[Dict[str, str]]:
    batch: Dict[str, str] = {}
    total_chars = 0
    for key, value in items.items():
        value_chars = len(value)
        if batch and (len(batch) >= TRANSLATE_BATCH_LIMIT or total_chars + value_chars > TRANSLATE_CHAR_LIMIT):
            yield batch
            batch = {}
            total_chars = 0
        batch[key] = value
        total_chars += value_chars
    if batch:
        yield batch


def call_llm_translate(items: Dict[str, str], target_lang: str) -> Dict[str, str]:
    if not items:
        return {}
    if MOCK_TRANSLATION:
        return {key: f"[{target_lang}] {value}" for key, value in items.items()}
    if not OPENAI_API_KEY:
        raise RuntimeError("缺少 OPENAI_API_KEY")

    system_prompt = (
        f"你是一个专业的翻译专家。请将用户提供的 JSON 格式文本翻译成 {target_lang}。"
        "保持 JSON 的键名不变，只翻译值。返回必须是合法 JSON 对象。"
    )

    response = httpx.post(
        OPENAI_BASE_URL,
        json={
            "model": MODEL_NAME,
            "messages": [
                {"role": "system", "content": system_prompt},
                {"role": "user", "content": json.dumps(items, ensure_ascii=False)},
            ],
            "temperature": 0.3,
            "response_format": {"type": "json_object"},
        },
        headers={
            "Content-Type": "application/json",
            "Authorization": f"Bearer {OPENAI_API_KEY}",
        },
        timeout=120,
    )
    response.raise_for_status()
    payload = response.json()
    content = payload["choices"][0]["message"]["content"]
    clean = content.replace("```json", "").replace("```", "").strip()
    parsed = json.loads(clean)
    return {key: str(value) for key, value in parsed.items()}


def translate_sheet(file_bytes: bytes, sheet_name: str, target_lang: str) -> Tuple[bytes, int]:
    original_map = extract_texts(file_bytes, sheet_name)
    if not original_map:
        return file_bytes, 0

    merged_translations: Dict[str, str] = {}
    for batch in chunk_items(original_map):
        translated = call_llm_translate(batch, target_lang)
        merged_translations.update(translated)

    updated_bytes = apply_translations(
        file_bytes=file_bytes,
        sheet_name=sheet_name,
        translations=merged_translations,
        original_texts_map=original_map,
    )
    return updated_bytes, len(original_map)


class AppHandler(BaseHTTPRequestHandler):
    server_version = "PythonExcelHTTP/1.0"

    def _set_common_headers(self, status_code: int, content_type: str) -> None:
        self.send_response(status_code)
        self.send_header("Content-Type", content_type)
        self.send_header("Access-Control-Allow-Origin", "*")
        self.send_header("Access-Control-Allow-Methods", "GET, POST, OPTIONS")
        self.send_header("Access-Control-Allow-Headers", "Content-Type, Authorization, x-admin-token")

    def _send_json(self, data: Dict, status_code: int = 200) -> None:
        body = json.dumps({"isBase64Encoded": False, "statusCode": status_code, "data": data}, ensure_ascii=False).encode("utf-8")
        self._set_common_headers(status_code, "application/json; charset=utf-8")
        self.send_header("Content-Length", str(len(body)))
        self.end_headers()
        self.wfile.write(body)

    def _read_json_body(self) -> Dict:
        length = int(self.headers.get("Content-Length", "0") or "0")
        if length <= 0:
            return {}
        raw = self.rfile.read(length)
        return json.loads(raw.decode("utf-8")) if raw else {}

    def do_OPTIONS(self) -> None:
        self._set_common_headers(204, "text/plain; charset=utf-8")
        self.end_headers()

    def do_GET(self) -> None:
        parsed = urlparse(self.path)
        if parsed.path == "/health":
            body = json.dumps({"ok": True}, ensure_ascii=False).encode("utf-8")
            self._set_common_headers(200, "application/json; charset=utf-8")
            self.send_header("Content-Length", str(len(body)))
            self.end_headers()
            self.wfile.write(body)
            return

        if parsed.path.startswith("/download/"):
            result_id = parsed.path.rsplit("/", 1)[-1]
            files = list(RESULT_DIR.glob(f"{result_id}.*"))
            if not files:
                self._send_json({"error": "结果文件不存在"}, 404)
                return
            path = files[0]
            query = parse_qs(parsed.query)
            filename = sanitize_filename((query.get("filename") or [path.name])[0])
            data = path.read_bytes()
            self._set_common_headers(200, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
            self.send_header("Content-Disposition", f'attachment; filename="{filename}"')
            self.send_header("Content-Length", str(len(data)))
            self.end_headers()
            self.wfile.write(data)
            return

        self._send_json({"error": "路径不存在"}, 404)

    def do_POST(self) -> None:
        parsed = urlparse(self.path)
        payload = self._read_json_body()

        try:
            if parsed.path == "/upload-file-function":
                action = payload.get("action")
                if action == "init":
                    upload_id = init_upload_session(
                        payload.get("file_name", ""),
                        payload.get("file_size", 0),
                        payload.get("content_type", "application/octet-stream"),
                    )
                    self._send_json({"upload_id": upload_id})
                    return
                if action == "append":
                    append_upload_chunk(
                        payload.get("upload_id", ""),
                        int(payload.get("chunk_index", 0)),
                        payload.get("chunk_base64", ""),
                    )
                    self._send_json({"ok": True})
                    return
                if action == "complete":
                    complete_upload_session(payload.get("upload_id", ""), int(payload.get("total_chunks", 0)))
                    self._send_json({"upload_id": payload.get("upload_id"), "total_chunks": int(payload.get("total_chunks", 0))})
                    return
                self._send_json({"error": "不支持的操作类型"}, 400)
                return

            if parsed.path == "/parse-function":
                upload_id = payload.get("upload_id")
                if not upload_id:
                    self._send_json({"error": "缺少 upload_id"}, 400)
                    return
                file_bytes, _meta = get_file_bytes(upload_id)
                sheet_names, previews_by_sheet, total_cells_by_sheet = build_sheet_previews(file_bytes)
                resolved_sheet = payload.get("sheet_name") if payload.get("sheet_name") in previews_by_sheet else (sheet_names[0] if sheet_names else "")
                total_cells = sum(total_cells_by_sheet.values())
                self._send_json(
                    {
                        "preview": previews_by_sheet.get(resolved_sheet, []),
                        "total_cells": total_cells,
                        "sheet_name": resolved_sheet,
                        "sheet_names": sheet_names,
                        "previews_by_sheet": previews_by_sheet,
                        "total_cells_by_sheet": total_cells_by_sheet,
                    }
                )
                return

            if parsed.path == "/translate-function":
                upload_id = payload.get("upload_id")
                target_lang = payload.get("target_lang")
                if not upload_id:
                    self._send_json({"error": "缺少 upload_id"}, 400)
                    return
                if not target_lang:
                    self._send_json({"error": "缺少 target_lang"}, 400)
                    return

                file_bytes, meta = get_file_bytes(upload_id)
                sheet_names = payload.get("sheet_names") or ([payload.get("sheet_name")] if payload.get("sheet_name") else [])
                if not sheet_names:
                    self._send_json({"error": "未找到可翻译的工作表"}, 400)
                    return

                current_bytes = file_bytes
                used_count = 0
                for sheet_name in sheet_names:
                    current_bytes, translated_count = translate_sheet(current_bytes, sheet_name, target_lang)
                    used_count += 1 if translated_count > 0 else 0

                result_id = uuid.uuid4().hex
                original_name = sanitize_filename(meta.get("file_name") or "translated.xlsx")
                stem, suffix = os.path.splitext(original_name)
                suffix = suffix or ".xlsx"
                output_name = sanitize_filename(f"{stem}-translated{suffix}")
                result_path = RESULT_DIR / f"{result_id}{suffix}"
                result_path.write_bytes(current_bytes)

                self._send_json(
                    {
                        "file_url": f"{API_PREFIX}/download/{result_id}?filename={quote(output_name)}",
                        "used_count": used_count or 1,
                    }
                )
                return

            self._send_json({"error": "路径不存在"}, 404)
        except Exception as exc:
            endpoint = parsed.path.strip("/") or "request"
            label = "解析失败" if endpoint == "parse-function" else "翻译失败" if endpoint == "translate-function" else "上传处理失败"
            self._send_json({"error": f"{label}: {exc}"}, 500)


if __name__ == "__main__":
    server = ThreadingHTTPServer(("0.0.0.0", 9000), AppHandler)
    server.serve_forever()
