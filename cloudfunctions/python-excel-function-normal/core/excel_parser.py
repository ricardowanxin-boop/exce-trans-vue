import os
import io
import xml.etree.ElementTree as ET
import zipfile
from typing import Dict, List, Optional, Tuple

import openpyxl

def _get_sheet_filename(zip_file: zipfile.ZipFile, sheet_name: str) -> Optional[str]:
    """获取指定Sheet对应的sheet xml文件名"""
    try:
        workbook_xml = zip_file.read('xl/workbook.xml')
        root = ET.fromstring(workbook_xml)
        ns = {'r': 'http://schemas.openxmlformats.org/officeDocument/2006/relationships'}

        r_id = None
        for sheet in root.findall('.//{*}sheet'):
            if sheet.get('name') == sheet_name:
                r_id = sheet.get(f'{{{ns["r"]}}}id')
                break

        if not r_id:
            return None

        rels_xml = zip_file.read('xl/_rels/workbook.xml.rels')
        rels_root = ET.fromstring(rels_xml)

        sheet_path = None
        for rel in rels_root.findall('.//{*}Relationship'):
            if rel.get('Id') == r_id:
                sheet_path = rel.get('Target')
                break

        if not sheet_path:
            return None

        if sheet_path.startswith('/'):
            sheet_path = sheet_path[1:]
        elif not sheet_path.startswith('xl/'):
            sheet_path = 'xl/' + sheet_path

        sheet_path = sheet_path.replace('\\', '/')
        return sheet_path if sheet_path in zip_file.namelist() else None

    except Exception as e:
        print(f"Error finding sheet file: {e}")
        return None

def _get_drawing_filename(zip_file: zipfile.ZipFile, sheet_name: str) -> Optional[str]:
    """获取指定Sheet对应的drawing.xml文件名"""
    try:
        sheet_path = _get_sheet_filename(zip_file, sheet_name)
        if not sheet_path:
            return None

        sheet_dir = os.path.dirname(sheet_path)
        sheet_fname = os.path.basename(sheet_path)
        rels_path = f"{sheet_dir}/_rels/{sheet_fname}.rels"

        if rels_path not in zip_file.namelist():
            return None

        sheet_rels_xml = zip_file.read(rels_path)
        sheet_rels_root = ET.fromstring(sheet_rels_xml)

        drawing_path = None
        for rel in sheet_rels_root.findall('.//{*}Relationship'):
            if rel.get('Type').endswith('/drawing'):
                drawing_path = rel.get('Target')
                break

        if not drawing_path:
            return None

        if drawing_path.startswith('../'):
            final_path = 'xl/' + drawing_path.replace('../', '')
        else:
            final_path = f"{sheet_dir}/{drawing_path}"

        final_path = final_path.replace('\\', '/')
        return final_path if final_path in zip_file.namelist() else None

    except Exception as e:
        print(f"Error finding drawing file: {e}")
        return None

def extract_shape_entries(file_bytes: bytes, sheet_name: str) -> List[Tuple[str, str]]:
    """提取形状文本并生成可展示的顺序编号。"""
    entries: List[Tuple[str, str]] = []
    try:
        with zipfile.ZipFile(io.BytesIO(file_bytes), 'r') as z:
            drawing_file = _get_drawing_filename(z, sheet_name)
            if not drawing_file:
                return []
                
            xml_content = z.read(drawing_file)
            root = ET.fromstring(xml_content)
            ns = {'a': 'http://schemas.openxmlformats.org/drawingml/2006/main'}
            paragraph_texts = set()

            for p_node in root.findall('.//a:p', ns):
                t_nodes = p_node.findall('.//a:t', ns)
                if not t_nodes:
                    continue
                full_text = "".join([t.text for t in t_nodes if t.text])
                cleaned = full_text.strip()
                if cleaned:
                    paragraph_texts.add(cleaned)
                    entries.append((f"S{len(entries) + 1}", cleaned))

            for t_node in root.findall('.//a:t', ns):
                if not t_node.text:
                    continue
                cleaned = t_node.text.strip()
                if not cleaned or cleaned in paragraph_texts:
                    continue
                entries.append((f"S{len(entries) + 1}", cleaned))
    except Exception as e:
        print(f"提取形状文本失败: {e}")

    return entries

def _replace_text_in_xml(
    xml_bytes: bytes,
    replacements: Dict[str, str],
    mode: str = 'sharedStrings'
) -> bytes:
    """在XML内容中替换文本，支持 sharedStrings / worksheet / drawing"""
    try:
        root = ET.fromstring(xml_bytes)
        namespaces = {
            'xdr': "http://schemas.openxmlformats.org/drawingml/2006/spreadsheetDrawing",
            'a': "http://schemas.openxmlformats.org/drawingml/2006/main",
            'r': "http://schemas.openxmlformats.org/officeDocument/2006/relationships",
            'main': "http://schemas.openxmlformats.org/spreadsheetml/2006/main"
        }
        for prefix, uri in namespaces.items():
            ET.register_namespace(prefix, uri)

        if mode != 'drawing':
             ET.register_namespace('', "http://schemas.openxmlformats.org/spreadsheetml/2006/main")

        modified = False

        if mode == 'drawing':
            ns = {'a': 'http://schemas.openxmlformats.org/drawingml/2006/main'}
            for p_node in root.findall('.//a:p', ns):
                t_nodes = p_node.findall('.//a:t', ns)
                if not t_nodes:
                    continue
                full_text = "".join([t.text for t in t_nodes if t.text])
                full_text_stripped = full_text.strip()

                if full_text in replacements or full_text_stripped in replacements:
                    target_translation = replacements.get(full_text) or replacements.get(full_text_stripped)
                    t_nodes[0].text = target_translation
                    for t in t_nodes[1:]:
                        t.text = ""
                    modified = True
                    continue

                for t_node in t_nodes:
                    if t_node.text:
                        node_text_stripped = t_node.text.strip()
                        if t_node.text in replacements:
                            t_node.text = replacements[t_node.text]
                            modified = True
                        elif node_text_stripped in replacements:
                            t_node.text = replacements[node_text_stripped]
                            modified = True

        elif mode == 'worksheet':
            ns = {'main': "http://schemas.openxmlformats.org/spreadsheetml/2006/main"}

            for cell_node in root.findall('.//main:c', ns):
                cell_type = cell_node.get('t')

                if cell_type == 'inlineStr':
                    t_nodes = cell_node.findall('.//main:is//main:t', ns)
                    if not t_nodes:
                        continue

                    full_text = "".join([t.text for t in t_nodes if t.text])
                    full_text_stripped = full_text.strip()

                    if full_text in replacements or full_text_stripped in replacements:
                        target_translation = replacements.get(full_text) or replacements.get(full_text_stripped)
                        t_nodes[0].text = target_translation
                        for t in t_nodes[1:]:
                            t.text = ""
                        modified = True
                        continue

                    for t_node in t_nodes:
                        if t_node.text:
                            text = t_node.text
                            text_stripped = text.strip()
                            if text in replacements:
                                t_node.text = replacements[text]
                                modified = True
                            elif text_stripped in replacements:
                                t_node.text = replacements[text_stripped]
                                modified = True

                elif cell_type == 'str':
                    v_node = cell_node.find('./main:v', ns)
                    if v_node is not None and v_node.text:
                        text = v_node.text
                        text_stripped = text.strip()
                        if text in replacements:
                            v_node.text = replacements[text]
                            modified = True
                        elif text_stripped in replacements:
                            v_node.text = replacements[text_stripped]
                            modified = True

        else:
            target_tag = f"{{{namespaces['main']}}}t"
            for elem in root.iter():
                if elem.tag == target_tag and elem.text:
                    elem_text_stripped = elem.text.strip()
                    if elem.text in replacements:
                        elem.text = replacements[elem.text]
                        modified = True
                    elif elem_text_stripped in replacements:
                        elem.text = replacements[elem_text_stripped]
                        modified = True
                        
        if modified:
            return ET.tostring(root, encoding='utf-8', method='xml', xml_declaration=True)
        return xml_bytes
        
    except Exception as e:
        print(f"XML替换失败: {e}")
        return xml_bytes

def apply_translations(
    file_bytes: bytes,
    sheet_name: str,
    translations: Dict[str, str],
    original_texts_map: Optional[Dict[str, str]] = None
) -> bytes:
    """
    将翻译结果写回Excel，采用"Surgeon"模式：
    直接修改原始文件的XML，不经过openpyxl的save，以最大程度保留文件结构（形状、样式等）。
    """
    if not original_texts_map:
        print("Warning: original_texts_map not provided, re-extracting...")
        original_texts_map = extract_texts(file_bytes, sheet_name)
    
    text_replacements = {}
    for coord, translated_text in translations.items():
        if coord in original_texts_map:
            original_text = original_texts_map[coord]
            if original_text and original_text != translated_text:
                text_replacements[original_text] = translated_text
                
    if not text_replacements:
        return file_bytes
        
    output_buffer = io.BytesIO()
    
    try:
        with zipfile.ZipFile(io.BytesIO(file_bytes), 'r') as zin:
            target_sheet_path = _get_sheet_filename(zin, sheet_name)
            target_drawing_path = _get_drawing_filename(zin, sheet_name)
            
            with zipfile.ZipFile(output_buffer, 'w') as zout:
                for item in zin.infolist():
                    data = zin.read(item.filename)
                    
                    if item.filename.endswith('sharedStrings.xml'):
                        data = _replace_text_in_xml(data, text_replacements, mode='sharedStrings')

                    elif target_sheet_path and item.filename == target_sheet_path:
                        data = _replace_text_in_xml(data, text_replacements, mode='worksheet')
                    
                    elif target_drawing_path and item.filename == target_drawing_path:
                        data = _replace_text_in_xml(data, text_replacements, mode='drawing')
                        
                    zout.writestr(item, data)
                    
        return output_buffer.getvalue()
        
    except Exception as e:
        print(f"Surgeon模式写入失败: {e}")
        raise e

def get_sheet_names(file_bytes: bytes) -> List[str]:
    """获取Excel文件中的所有Sheet名称"""
    wb = openpyxl.load_workbook(io.BytesIO(file_bytes), read_only=True)
    return wb.sheetnames

def extract_texts(
    file_bytes: bytes,
    sheet_name: str,
    ignore_formulas: bool = True,
    ignore_numbers: bool = True,
    ignore_header_rows: int = 0
) -> Dict[str, str]:
    """
    提取待翻译的文本
    返回字典格式: {"A1": "原文", "B2": "原文"}
    """
    wb = openpyxl.load_workbook(io.BytesIO(file_bytes), data_only=False)
    
    if sheet_name not in wb.sheetnames:
        raise ValueError(f"工作表 {sheet_name} 不存在")
        
    ws = wb[sheet_name]
    texts_to_translate = {}
    
    for row_idx, row in enumerate(ws.iter_rows(), start=1):
        if row_idx <= ignore_header_rows:
            continue
            
        for cell in row:
            if cell.value is None:
                continue
                
            if ignore_formulas and cell.data_type == 'f':
                continue
                
            if ignore_numbers and isinstance(cell.value, (int, float)):
                continue
                
            text = str(cell.value).strip()
            if not text:
                continue
                
            texts_to_translate[cell.coordinate] = text
            
    for shape_coordinate, shape_text in extract_shape_entries(file_bytes, sheet_name):
        texts_to_translate[shape_coordinate] = shape_text
    
    return texts_to_translate
