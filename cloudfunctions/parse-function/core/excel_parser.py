import io
import openpyxl
import zipfile
import xml.etree.ElementTree as ET
import os
from typing import Dict, List, Tuple, Optional

def _get_drawing_filename(zip_file: zipfile.ZipFile, sheet_name: str) -> Optional[str]:
    """获取指定Sheet对应的drawing.xml文件名"""
    try:
        # 1. 解析 workbook.xml 找到 sheet 的 rId
        workbook_xml = zip_file.read('xl/workbook.xml')
        root = ET.fromstring(workbook_xml)
        # 命名空间处理，忽略默认命名空间前缀匹配
        ns = {'r': 'http://schemas.openxmlformats.org/officeDocument/2006/relationships'}
        
        r_id = None
        # 查找 sheet 节点，忽略命名空间前缀
        for sheet in root.findall('.//{*}sheet'):
            if sheet.get('name') == sheet_name:
                r_id = sheet.get(f'{{{ns["r"]}}}id')
                break
        
        if not r_id:
            return None
            
        # 2. 解析 workbook.xml.rels 找到 sheet 文件的路径
        rels_xml = zip_file.read('xl/_rels/workbook.xml.rels')
        rels_root = ET.fromstring(rels_xml)
        
        sheet_path = None
        for rel in rels_root.findall('.//{*}Relationship'):
            if rel.get('Id') == r_id:
                sheet_path = rel.get('Target')
                break
                
        if not sheet_path:
            return None
            
        # 修正路径，确保是相对于 zip 根目录的路径
        if sheet_path.startswith('/'):
            sheet_path = sheet_path[1:]
        elif not sheet_path.startswith('xl/'):
            sheet_path = 'xl/' + sheet_path
            
        # 3. 解析 sheetN.xml.rels 找到 drawing 文件的路径
        sheet_dir = os.path.dirname(sheet_path)
        sheet_fname = os.path.basename(sheet_path)
        rels_path = f"{sheet_dir}/_rels/{sheet_fname}.rels"
        
        if rels_path not in zip_file.namelist():
            return None
            
        sheet_rels_xml = zip_file.read(rels_path)
        sheet_rels_root = ET.fromstring(sheet_rels_xml)
        
        drawing_path = None
        for rel in sheet_rels_root.findall('.//{*}Relationship'):
            if rel.get('Type').endswith('/drawing'):
                drawing_path = rel.get('Target')
                break
                
        if not drawing_path:
            return None
            
        # 修正 drawing 路径
        # 通常 drawing_path 是 "../drawings/drawing1.xml" 形式
        if drawing_path.startswith('../'):
            final_path = 'xl/' + drawing_path.replace('../', '')
        else:
            final_path = f"{sheet_dir}/{drawing_path}"
            
        # 规范化路径分隔符
        final_path = final_path.replace('\\', '/')
        
        return final_path if final_path in zip_file.namelist() else None
        
    except Exception as e:
        print(f"Error finding drawing file: {e}")
        return None

def _extract_shape_texts(file_bytes: bytes, sheet_name: str) -> Dict[str, str]:
    """从形状中提取文本"""
    texts = {}
    try:
        with zipfile.ZipFile(io.BytesIO(file_bytes), 'r') as z:
            drawing_file = _get_drawing_filename(z, sheet_name)
            if not drawing_file:
                return {}
                
            xml_content = z.read(drawing_file)
            root = ET.fromstring(xml_content)
            # drawingml main namespace
            ns = {'a': 'http://schemas.openxmlformats.org/drawingml/2006/main'}
            
            # 按段落提取，以保证长句完整性
            for p_node in root.findall('.//a:p', ns):
                t_nodes = p_node.findall('.//a:t', ns)
                if not t_nodes:
                    continue
                # 拼接整个段落的内容
                full_text = "".join([t.text for t in t_nodes if t.text])
                if full_text.strip():
                    texts[f"SHAPE_P||{full_text}"] = full_text.strip()
                    
            # 同时也保留原来的按节点提取，作为双保险
            for t_node in root.findall('.//a:t', ns):
                if t_node.text and t_node.text.strip():
                    # 使用特殊前缀标识形状文本，键即为原始内容，方便后续替换
                    texts[f"SHAPE||{t_node.text}"] = t_node.text.strip()
    except Exception as e:
        print(f"提取形状文本失败: {e}")
        
    return texts

def _replace_text_in_xml(xml_bytes: bytes, replacements: Dict[str, str], is_drawing: bool = False) -> bytes:
    """在XML内容中替换文本，支持sharedStrings和drawing"""
    try:
        root = ET.fromstring(xml_bytes)
        # 注册所有可能的命名空间
        namespaces = {
            'xdr': "http://schemas.openxmlformats.org/drawingml/2006/spreadsheetDrawing",
            'a': "http://schemas.openxmlformats.org/drawingml/2006/main",
            'r': "http://schemas.openxmlformats.org/officeDocument/2006/relationships",
            'main': "http://schemas.openxmlformats.org/spreadsheetml/2006/main"
        }
        for prefix, uri in namespaces.items():
            ET.register_namespace(prefix, uri)
            
        # 默认命名空间处理
        if not is_drawing:
             ET.register_namespace('', "http://schemas.openxmlformats.org/spreadsheetml/2006/main")
        
        modified = False
        
        # 查找所有的文本节点
        # sharedStrings: <t>Content</t>
        # drawing: <a:t>Content</a:t>
        
        if is_drawing:
            # Drawing XML logic
            ns = {'a': 'http://schemas.openxmlformats.org/drawingml/2006/main'}
            
            # 找到所有的段落 <a:p>
            for p_node in root.findall('.//a:p', ns):
                # 获取该段落下的所有文本节点
                t_nodes = p_node.findall('.//a:t', ns)
                if not t_nodes:
                    continue
                    
                # 尝试拼接文本
                full_text = "".join([t.text for t in t_nodes if t.text])
                
                # 如果拼接后的完整文本在替换列表中
                full_text_stripped = full_text.strip()
                
                if full_text in replacements or full_text_stripped in replacements:
                    target_translation = replacements.get(full_text) or replacements.get(full_text_stripped)
                    t_nodes[0].text = target_translation
                    # 清空其他节点
                    for t in t_nodes[1:]:
                        t.text = ""
                    modified = True
                    continue
                
                # 新增逻辑：如果段落拼接没有匹配上，尝试将 replacements 里的所有长文本作为子串去匹配当前段落
                for t_node in t_nodes:
                    if t_node.text:
                        node_text_stripped = t_node.text.strip()
                        if t_node.text in replacements:
                            t_node.text = replacements[t_node.text]
                            modified = True
                        elif node_text_stripped in replacements:
                            t_node.text = replacements[node_text_stripped]
                            modified = True
                        else:
                            if len(node_text_stripped) > 3:
                                for orig_k, trans_v in replacements.items():
                                    if node_text_stripped in orig_k:
                                        pass
            
        else:
            # SharedStrings XML logic
            target_tag = f"{{{namespaces['main']}}}t"
            for elem in root.iter():
                if elem.tag == target_tag and elem.text:
                    elem_text_stripped = elem.text.strip()
                    if elem.text in replacements:
                        elem.text = replacements[elem.text]
                        modified = True
                    elif elem_text_stripped in replacements:
                        elem.text = replacements[elem_text_stripped]
                        modified = True
                        
        if modified:
            return ET.tostring(root, encoding='utf-8', method='xml', xml_declaration=True)
        return xml_bytes
        
    except Exception as e:
        print(f"XML替换失败: {e}")
        return xml_bytes

def apply_translations(
    file_bytes: bytes,
    sheet_name: str,
    translations: Dict[str, str],
    original_texts_map: Optional[Dict[str, str]] = None
) -> bytes:
    """
    将翻译结果写回Excel，采用"Surgeon"模式：
    直接修改原始文件的XML，不经过openpyxl的save，以最大程度保留文件结构（形状、样式等）。
    """
    if not original_texts_map:
        print("Warning: original_texts_map not provided, re-extracting...")
        original_texts_map = extract_texts(file_bytes, sheet_name)
    
    text_replacements = {}
    for coord, translated_text in translations.items():
        if coord in original_texts_map:
            original_text = original_texts_map[coord]
            if original_text and original_text != translated_text:
                text_replacements[original_text] = translated_text
                
    if not text_replacements:
        return file_bytes
        
    output_buffer = io.BytesIO()
    
    try:
        with zipfile.ZipFile(io.BytesIO(file_bytes), 'r') as zin:
            target_drawing_path = _get_drawing_filename(zin, sheet_name)
            
            with zipfile.ZipFile(output_buffer, 'w') as zout:
                for item in zin.infolist():
                    data = zin.read(item.filename)
                    
                    if item.filename.endswith('sharedStrings.xml'):
                        data = _replace_text_in_xml(data, text_replacements, is_drawing=False)
                    
                    elif target_drawing_path and item.filename == target_drawing_path:
                        data = _replace_text_in_xml(data, text_replacements, is_drawing=True)
                        
                    zout.writestr(item, data)
                    
        return output_buffer.getvalue()
        
    except Exception as e:
        print(f"Surgeon模式写入失败: {e}")
        raise e

def get_sheet_names(file_bytes: bytes) -> List[str]:
    """获取Excel文件中的所有Sheet名称"""
    wb = openpyxl.load_workbook(io.BytesIO(file_bytes), read_only=True)
    return wb.sheetnames

def extract_texts(
    file_bytes: bytes,
    sheet_name: str,
    ignore_formulas: bool = True,
    ignore_numbers: bool = True,
    ignore_header_rows: int = 0
) -> Dict[str, str]:
    """
    提取待翻译的文本
    返回字典格式: {"A1": "原文", "B2": "原文"}
    """
    wb = openpyxl.load_workbook(io.BytesIO(file_bytes), data_only=False)
    
    if sheet_name not in wb.sheetnames:
        raise ValueError(f"工作表 {sheet_name} 不存在")
        
    ws = wb[sheet_name]
    texts_to_translate = {}
    
    for row_idx, row in enumerate(ws.iter_rows(), start=1):
        if row_idx <= ignore_header_rows:
            continue
            
        for cell in row:
            if cell.value is None:
                continue
                
            if ignore_formulas and cell.data_type == 'f':
                continue
                
            if ignore_numbers and isinstance(cell.value, (int, float)):
                continue
                
            text = str(cell.value).strip()
            if not text:
                continue
                
            texts_to_translate[cell.coordinate] = text
            
    shape_texts = _extract_shape_texts(file_bytes, sheet_name)
    texts_to_translate.update(shape_texts)
    
    return texts_to_translate
