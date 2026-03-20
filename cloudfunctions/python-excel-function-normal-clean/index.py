from __future__ import annotations

import base64
import io
import json
import os
import re
import sys
import uuid
from pathlib import Path
from typing import Dict, Iterable, List, Tuple
from urllib.parse import quote, unquote

sys.path.insert(0, str(Path(__file__).resolve().parent / "third_party"))

import httpx
import openpyxl

from core.excel_parser import apply_translations, extract_shape_entries, extract_texts


BASE_DIR = Path(__file__).resolve().parent
IS_CLOUDBASE_RUNTIME = str(BASE_DIR).startswith("/var/")
RUNTIME_DIR = Path(
    os.getenv("PY_RUNTIME_DIR")
    or ("/tmp/python-excel-function-runtime" if IS_CLOUDBASE_RUNTIME else str(BASE_DIR / ".runtime"))
)
UPLOAD_DIR = RUNTIME_DIR / "uploads"
RESULT_DIR = RUNTIME_DIR / "results"
META_DIR = RUNTIME_DIR / "meta"
API_PREFIX = os.getenv("PY_API_PREFIX", "/python-api")
PREVIEW_LIMIT_PER_SHEET = 30
TRANSLATE_BATCH_LIMIT = 80
TRANSLATE_CHAR_LIMIT = 12000
RESULT_CHUNK_SIZE = 256 * 1024
OPENAI_API_KEY = os.getenv("OPENAI_API_KEY", "")
OPENAI_BASE_URL = os.getenv("OPENAI_BASE_URL", "https://api-inference.modelscope.cn/v1/chat/completions")
MODEL_NAME = os.getenv("MODEL_NAME", "Qwen/Qwen3-30B-A3B-Instruct-2507")
MOCK_TRANSLATION = os.getenv("PY_MOCK_TRANSLATION", "false").lower() == "true"

for directory in (UPLOAD_DIR, RESULT_DIR, META_DIR):
    directory.mkdir(parents=True, exist_ok=True)


def make_json_response(data: Dict, status_code: int = 200, headers: Dict[str, str] | None = None) -> Dict:
    merged_headers = {
        "Content-Type": "application/json; charset=utf-8",
        "Access-Control-Allow-Origin": "*",
        "Access-Control-Allow-Methods": "GET,POST,OPTIONS",
        "Access-Control-Allow-Headers": "Content-Type,Authorization,x-admin-token",
    }
    if headers:
        merged_headers.update(headers)

    return {
        "statusCode": status_code,
        "headers": merged_headers,
        "body": json.dumps({"isBase64Encoded": False, "statusCode": status_code, "data": data}, ensure_ascii=False),
        "isBase64Encoded": False,
    }


def make_binary_response(data: bytes, filename: str, content_type: str) -> Dict:
    return {
        "statusCode": 200,
        "headers": {
            "Content-Type": content_type,
            "Content-Disposition": f'attachment; filename="{filename}"',
            "Access-Control-Allow-Origin": "*",
            "Access-Control-Allow-Methods": "GET,POST,OPTIONS",
            "Access-Control-Allow-Headers": "Content-Type,Authorization,x-admin-token",
        },
        "body": base64.b64encode(data).decode("ascii"),
        "isBase64Encoded": True,
    }


def sanitize_filename(name: str) -> str:
    clean = re.sub(r"[^\w.\-()\u4e00-\u9fff]+", "_", name or "translated.xlsx").strip("._")
    return clean or "translated.xlsx"


def meta_path(upload_id: str) -> Path:
    return META_DIR / f"{upload_id}.json"


def upload_path(upload_id: str) -> Path:
    return UPLOAD_DIR / f"{upload_id}.bin"


def load_meta(upload_id: str) -> Dict:
    path = meta_path(upload_id)
    if not path.exists():
        raise FileNotFoundError("上传会话不存在")
    return json.loads(path.read_text(encoding="utf-8"))


def save_meta(upload_id: str, payload: Dict) -> None:
    meta_path(upload_id).write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8")


def init_upload_session(file_name: str, file_size: int, content_type: str) -> str:
    upload_id = f"upload_{uuid.uuid4().hex}"
    payload = {
        "upload_id": upload_id,
        "file_name": sanitize_filename(file_name),
        "file_size": max(int(file_size or 0), 0),
        "content_type": content_type or "application/octet-stream",
        "status": "uploading",
        "next_chunk_index": 0,
        "received_chunks": 0,
        "total_chunks": 0,
    }
    upload_path(upload_id).write_bytes(b"")
    save_meta(upload_id, payload)
    return upload_id


def append_upload_chunk(upload_id: str, chunk_index: int, chunk_base64: str) -> None:
    meta = load_meta(upload_id)
    if meta.get("status") != "uploading":
        raise ValueError("上传会话已关闭")
    expected_index = int(meta.get("next_chunk_index", 0))
    if chunk_index != expected_index:
        raise ValueError(f"分片顺序错误，期望 {expected_index}，收到 {chunk_index}")

    chunk_bytes = base64.b64decode(chunk_base64)
    with upload_path(upload_id).open("ab") as file:
        file.write(chunk_bytes)

    meta["next_chunk_index"] = expected_index + 1
    meta["received_chunks"] = int(meta.get("received_chunks", 0)) + 1
    save_meta(upload_id, meta)


def complete_upload_session(upload_id: str, total_chunks: int) -> None:
    meta = load_meta(upload_id)
    if int(meta.get("received_chunks", 0)) != total_chunks:
        raise ValueError("上传分片数量不完整")
    meta["status"] = "completed"
    meta["total_chunks"] = total_chunks
    save_meta(upload_id, meta)


def get_file_bytes(upload_id: str) -> Tuple[bytes, Dict]:
    meta = load_meta(upload_id)
    if meta.get("status") != "completed":
        raise ValueError("文件仍在上传中，请稍后重试")
    path = upload_path(upload_id)
    if not path.exists():
        raise FileNotFoundError("上传文件不存在")
    return path.read_bytes(), meta


def find_result_file(result_id: str) -> Path:
    files = list(RESULT_DIR.glob(f"{result_id}.*"))
    if not files:
        raise FileNotFoundError("结果文件不存在")
    return files[0]


def build_sheet_previews(file_bytes: bytes) -> Tuple[List[str], Dict[str, List[Dict[str, str]]], Dict[str, int]]:
    workbook = openpyxl.load_workbook(io.BytesIO(file_bytes), read_only=True, data_only=False)
    sheet_names = list(workbook.sheetnames)
    previews_by_sheet: Dict[str, List[Dict[str, str]]] = {}
    total_cells_by_sheet: Dict[str, int] = {}

    for sheet_name in sheet_names:
        worksheet = workbook[sheet_name]
        shape_preview = [
            {"coordinate": coordinate, "text": text, "source": "shape"}
            for coordinate, text in extract_shape_entries(file_bytes, sheet_name)
        ]
        preview: List[Dict[str, str]] = shape_preview[:PREVIEW_LIMIT_PER_SHEET]
        total = len(shape_preview)

        for row in worksheet.iter_rows():
            for cell in row:
                value = cell.value
                if value is None:
                    continue
                if cell.data_type == "f":
                    continue
                if isinstance(value, (int, float)):
                    continue
                text = str(value).strip()
                if not text:
                    continue
                total += 1
                if len(preview) < PREVIEW_LIMIT_PER_SHEET:
                    preview.append({"coordinate": cell.coordinate, "text": text, "source": "cell"})

        previews_by_sheet[sheet_name] = preview
        total_cells_by_sheet[sheet_name] = total

    workbook.close()
    return sheet_names, previews_by_sheet, total_cells_by_sheet


def chunk_items(items: Dict[str, str]) -> Iterable[Dict[str, str]]:
    batch: Dict[str, str] = {}
    total_chars = 0
    for key, value in items.items():
        value_chars = len(value)
        if batch and (len(batch) >= TRANSLATE_BATCH_LIMIT or total_chars + value_chars > TRANSLATE_CHAR_LIMIT):
            yield batch
            batch = {}
            total_chars = 0
        batch[key] = value
        total_chars += value_chars
    if batch:
        yield batch


def call_llm_translate(items: Dict[str, str], target_lang: str) -> Dict[str, str]:
    if not items:
        return {}
    if MOCK_TRANSLATION:
        return {key: f"[{target_lang}] {value}" for key, value in items.items()}
    if not OPENAI_API_KEY:
        raise RuntimeError("缺少 OPENAI_API_KEY")

    system_prompt = (
        f"你是一个专业的翻译专家。请将用户提供的 JSON 格式文本翻译成 {target_lang}。"
        "保持 JSON 的键名不变，只翻译值。返回必须是合法 JSON 对象。"
    )

    response = httpx.post(
        OPENAI_BASE_URL,
        json={
            "model": MODEL_NAME,
            "messages": [
                {"role": "system", "content": system_prompt},
                {"role": "user", "content": json.dumps(items, ensure_ascii=False)},
            ],
            "temperature": 0.3,
            "response_format": {"type": "json_object"},
        },
        headers={
            "Content-Type": "application/json",
            "Authorization": f"Bearer {OPENAI_API_KEY}",
        },
        timeout=120,
    )
    response.raise_for_status()
    payload = response.json()
    content = payload["choices"][0]["message"]["content"]
    clean = content.replace("```json", "").replace("```", "").strip()
    parsed = json.loads(clean)
    return {key: str(value) for key, value in parsed.items()}


def translate_sheet(file_bytes: bytes, sheet_name: str, target_lang: str) -> Tuple[bytes, int]:
    original_map = extract_texts(file_bytes, sheet_name)
    if not original_map:
        return file_bytes, 0

    merged_translations: Dict[str, str] = {}
    for batch in chunk_items(original_map):
        translated = call_llm_translate(batch, target_lang)
        merged_translations.update(translated)

    updated_bytes = apply_translations(
        file_bytes=file_bytes,
        sheet_name=sheet_name,
        translations=merged_translations,
        original_texts_map=original_map,
    )
    return updated_bytes, len(original_map)


def parse_event_body(event: Dict) -> Dict:
    body = event.get("body")
    if body is None:
        return event if isinstance(event, dict) else {}

    if isinstance(body, dict):
        return body

    if event.get("isBase64Encoded"):
        body = base64.b64decode(body).decode("utf-8")

    if isinstance(body, str) and body.strip():
        return json.loads(body)

    return {}


def resolve_path(event: Dict) -> str:
    path = event.get("path") or event.get("requestContext", {}).get("path") or "/"
    if not isinstance(path, str):
        return "/"
    if API_PREFIX and path == API_PREFIX:
        return "/"
    if API_PREFIX and path.startswith(f"{API_PREFIX}/"):
        return path[len(API_PREFIX):]
    return path


def handle_upload(payload: Dict) -> Dict:
    action = payload.get("action")
    if action == "init":
        upload_id = init_upload_session(
            payload.get("file_name", ""),
            payload.get("file_size", 0),
            payload.get("content_type", "application/octet-stream"),
        )
        return make_json_response({"upload_id": upload_id})

    if action == "append":
        append_upload_chunk(
            payload.get("upload_id", ""),
            int(payload.get("chunk_index", 0)),
            payload.get("chunk_base64", ""),
        )
        return make_json_response({"ok": True})

    if action == "complete":
        complete_upload_session(payload.get("upload_id", ""), int(payload.get("total_chunks", 0)))
        return make_json_response({"upload_id": payload.get("upload_id"), "total_chunks": int(payload.get("total_chunks", 0))})

    return make_json_response({"error": "不支持的操作类型"}, 400)


def handle_parse(payload: Dict) -> Dict:
    upload_id = payload.get("upload_id")
    if not upload_id:
        return make_json_response({"error": "缺少 upload_id"}, 400)

    file_bytes, _meta = get_file_bytes(upload_id)
    sheet_names, previews_by_sheet, total_cells_by_sheet = build_sheet_previews(file_bytes)
    resolved_sheet = payload.get("sheet_name") if payload.get("sheet_name") in previews_by_sheet else (sheet_names[0] if sheet_names else "")
    total_cells = sum(total_cells_by_sheet.values())
    return make_json_response(
        {
            "preview": previews_by_sheet.get(resolved_sheet, []),
            "total_cells": total_cells,
            "sheet_name": resolved_sheet,
            "sheet_names": sheet_names,
            "previews_by_sheet": previews_by_sheet,
            "total_cells_by_sheet": total_cells_by_sheet,
        }
    )


def handle_translate(payload: Dict) -> Dict:
    upload_id = payload.get("upload_id")
    target_lang = payload.get("target_lang")
    if not upload_id:
        return make_json_response({"error": "缺少 upload_id"}, 400)
    if not target_lang:
        return make_json_response({"error": "缺少 target_lang"}, 400)

    file_bytes, meta = get_file_bytes(upload_id)
    sheet_names = payload.get("sheet_names") or ([payload.get("sheet_name")] if payload.get("sheet_name") else [])
    if not sheet_names:
        return make_json_response({"error": "未找到可翻译的工作表"}, 400)

    current_bytes = file_bytes
    used_count = 0
    for sheet_name in sheet_names:
        current_bytes, translated_count = translate_sheet(current_bytes, sheet_name, target_lang)
        used_count += 1 if translated_count > 0 else 0

    result_id = uuid.uuid4().hex
    original_name = sanitize_filename(meta.get("file_name") or "translated.xlsx")
    stem, suffix = os.path.splitext(original_name)
    suffix = suffix or ".xlsx"
    output_name = sanitize_filename(f"{stem}-translated{suffix}")
    result_path = RESULT_DIR / f"{result_id}{suffix}"
    result_path.write_bytes(current_bytes)
    total_chunks = max(1, (len(current_bytes) + RESULT_CHUNK_SIZE - 1) // RESULT_CHUNK_SIZE)

    return make_json_response(
        {
            "file_url": f"{API_PREFIX}/download/{result_id}?filename={quote(output_name)}",
            "result_id": result_id,
            "file_name": output_name,
            "file_size": len(current_bytes),
            "total_chunks": total_chunks,
            "used_count": used_count or 1,
        }
    )


def handle_download_result_info(payload: Dict) -> Dict:
    result_id = str(payload.get("result_id") or "").strip()
    if not result_id:
        return make_json_response({"error": "缺少 result_id"}, 400)

    result_path = find_result_file(result_id)
    file_size = result_path.stat().st_size
    total_chunks = max(1, (file_size + RESULT_CHUNK_SIZE - 1) // RESULT_CHUNK_SIZE)
    return make_json_response(
        {
            "result_id": result_id,
            "file_name": payload.get("file_name") or result_path.name,
            "file_size": file_size,
            "total_chunks": total_chunks,
            "chunk_size": RESULT_CHUNK_SIZE,
        }
    )


def handle_download_result_chunk(payload: Dict) -> Dict:
    result_id = str(payload.get("result_id") or "").strip()
    if not result_id:
        return make_json_response({"error": "缺少 result_id"}, 400)

    chunk_index = int(payload.get("chunk_index") or 0)
    if chunk_index < 0:
        return make_json_response({"error": "chunk_index 不合法"}, 400)

    result_path = find_result_file(result_id)
    file_size = result_path.stat().st_size
    total_chunks = max(1, (file_size + RESULT_CHUNK_SIZE - 1) // RESULT_CHUNK_SIZE)
    if chunk_index >= total_chunks:
        return make_json_response({"error": "chunk_index 超出范围"}, 400)

    start = chunk_index * RESULT_CHUNK_SIZE
    end = min(file_size, start + RESULT_CHUNK_SIZE)
    with result_path.open("rb") as result_file:
        result_file.seek(start)
        chunk_bytes = result_file.read(end - start)

    return make_json_response(
        {
            "result_id": result_id,
            "chunk_index": chunk_index,
            "total_chunks": total_chunks,
            "chunk_base64": base64.b64encode(chunk_bytes).decode("ascii"),
        }
    )


def handle_download(path: str, event: Dict) -> Dict:
    result_id = path.rsplit("/", 1)[-1]
    files = list(RESULT_DIR.glob(f"{result_id}.*"))
    if not files:
        return make_json_response({"error": "结果文件不存在"}, 404)

    query = event.get("queryStringParameters") or {}
    filename = sanitize_filename(unquote((query.get("filename") or files[0].name)))
    return make_binary_response(
        data=files[0].read_bytes(),
        filename=filename,
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )


def is_http_event(event: Dict) -> bool:
    return isinstance(event, dict) and (
        "httpMethod" in event or "requestContext" in event or ("path" in event and "headers" in event)
    )


def unwrap_http_response(response: Dict) -> Dict:
    status_code = int(response.get("statusCode") or 500)
    body = response.get("body")
    parsed = json.loads(body) if isinstance(body, str) and body else {}
    data = parsed.get("data", parsed) if isinstance(parsed, dict) else parsed
    if status_code >= 400:
        error_message = ""
        if isinstance(data, dict):
            error_message = str(data.get("error") or "")
        if not error_message:
            error_message = f"请求失败 ({status_code})"
        return {"error": error_message, "statusCode": status_code}
    return {"data": data if data is not None else {}, "statusCode": status_code}


def dispatch_request(method: str, path: str, event: Dict, payload: Dict) -> Dict:
    if method == "OPTIONS":
        return {
            "statusCode": 204,
            "headers": {
                "Access-Control-Allow-Origin": "*",
                "Access-Control-Allow-Methods": "GET,POST,OPTIONS",
                "Access-Control-Allow-Headers": "Content-Type,Authorization,x-admin-token",
            },
            "body": "",
            "isBase64Encoded": False,
        }

    if path == "/health":
        return {
            "statusCode": 200,
            "headers": {
                "Content-Type": "application/json; charset=utf-8",
                "Access-Control-Allow-Origin": "*",
            },
            "body": json.dumps({"ok": True}, ensure_ascii=False),
            "isBase64Encoded": False,
        }

    if method == "GET" and path.startswith("/download/"):
        return handle_download(path, event)

    if method == "POST" and path == "/upload-file-function":
        return handle_upload(payload)

    if method == "POST" and path == "/parse-function":
        return handle_parse(payload)

    if method == "POST" and path == "/translate-function":
        return handle_translate(payload)

    if method == "POST" and path == "/download-result-info":
        return handle_download_result_info(payload)

    if method == "POST" and path == "/download-result-chunk":
        return handle_download_result_chunk(payload)

    return make_json_response({"error": f"未匹配路径: {path}"}, 404)


def main_handler(event, context):
    try:
        payload = parse_event_body(event)
        if is_http_event(event):
            method = (event.get("httpMethod") or event.get("requestContext", {}).get("httpMethod") or "GET").upper()
            path = resolve_path(event)
            return dispatch_request(method, path, event, payload)

        route = str(payload.get("__route") or payload.get("route") or "").strip() or "/"
        method = str(payload.get("__method") or "POST").upper()
        if not route.startswith("/"):
            route = f"/{route}"

        direct_payload = dict(payload)
        direct_payload.pop("__route", None)
        direct_payload.pop("route", None)
        direct_payload.pop("__method", None)

        response = dispatch_request(method, route, {}, direct_payload)
        if response.get("isBase64Encoded"):
            return {
                "error": "普通云函数直调不支持二进制直出，请改用 download-result-info / chunk",
                "statusCode": int(response.get("statusCode") or 500),
            }
        return unwrap_http_response(response)
    except Exception as exc:
        if is_http_event(event):
            return make_json_response({"error": f"Server Error: {exc}"}, 500)
        return {"error": f"Server Error: {exc}", "statusCode": 500}
